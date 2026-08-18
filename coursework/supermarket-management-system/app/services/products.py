import csv
import io

from app import db
from app.models import Category, Inventory, Product
from app.services.inventory import set_inventory_quantity


def import_products_from_csv(file_content):
    """
    从 CSV 导入商品

    Args:
        file_content: CSV 文件内容（字节）

    Returns:
        tuple: (success_count: int, error_count: int, errors: list)
    """
    success_count = 0
    error_count = 0
    errors = []

    try:
        text_content = file_content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text_content))

        required_fields = ['product_code', 'product_name', 'selling_price']
        if not all(field in reader.fieldnames for field in required_fields):
            return 0, 0, [f'CSV 文件缺少必要字段: {", ".join(required_fields)}']

        for row_num, row in enumerate(reader, start=2):
            try:
                product_code = row.get('product_code', '').strip()
                product_name = row.get('product_name', '').strip()
                selling_price = row.get('selling_price', '').strip()

                if not product_code or not product_name or not selling_price:
                    errors.append(f'第 {row_num} 行：缺少必要字段')
                    error_count += 1
                    continue

                if Product.query.filter_by(product_code=product_code).first():
                    errors.append(f'第 {row_num} 行：商品编码 {product_code} 已存在')
                    error_count += 1
                    continue

                barcode = row.get('barcode', '').strip() or None
                if barcode and Product.query.filter_by(barcode=barcode).first():
                    errors.append(f'第 {row_num} 行：条形码 {barcode} 已存在')
                    error_count += 1
                    continue

                product = Product(
                    product_code=product_code,
                    barcode=barcode,
                    product_name=product_name,
                    category_id=int(row['category_id']) if row.get('category_id', '').strip() else None,
                    unit=row.get('unit', '件').strip() or '件',
                    purchase_price=float(row.get('purchase_price', 0) or 0),
                    selling_price=float(selling_price),
                    min_stock=int(row.get('min_stock', 10) or 10),
                    status=1,
                )

                db.session.add(product)
                db.session.flush()

                quantity = int(row.get('quantity', 0) or 0)
                set_inventory_quantity(
                    product.product_id,
                    quantity,
                    change_type='import',
                    reason='CSV导入初始库存',
                )

                success_count += 1

            except ValueError as e:
                errors.append(f'第 {row_num} 行：数据格式错误 - {str(e)}')
                error_count += 1
            except Exception as e:
                errors.append(f'第 {row_num} 行：{str(e)}')
                error_count += 1

        db.session.commit()
        return success_count, error_count, errors

    except Exception as e:
        db.session.rollback()
        return 0, 0, [f'导入失败：{str(e)}']


def import_products_from_excel(file_content):
    """
    从 Excel 导入商品

    Args:
        file_content: Excel 文件内容（字节）

    Returns:
        tuple: (success_count: int, error_count: int, errors: list)
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        required_fields = ['product_code', 'product_name', 'selling_price']
        if not all(field in headers for field in required_fields):
            return 0, 0, [f'Excel 文件缺少必要字段: {", ".join(required_fields)}']

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            try:
                row_data = dict(zip(headers, row))

                product_code = str(row_data.get('product_code', '')).strip()
                product_name = str(row_data.get('product_name', '')).strip()
                selling_price = row_data.get('selling_price')

                if not product_code or not product_name or selling_price is None:
                    errors.append(f'第 {row_num} 行：缺少必要字段')
                    error_count += 1
                    continue

                if Product.query.filter_by(product_code=product_code).first():
                    errors.append(f'第 {row_num} 行：商品编码 {product_code} 已存在')
                    error_count += 1
                    continue

                barcode = str(row_data.get('barcode', '')).strip() if row_data.get('barcode') else None
                if barcode and Product.query.filter_by(barcode=barcode).first():
                    errors.append(f'第 {row_num} 行：条形码 {barcode} 已存在')
                    error_count += 1
                    continue

                product = Product(
                    product_code=product_code,
                    barcode=barcode,
                    product_name=product_name,
                    category_id=int(row_data['category_id']) if row_data.get('category_id') else None,
                    unit=str(row_data.get('unit', '件')).strip() or '件',
                    purchase_price=float(row_data.get('purchase_price', 0) or 0),
                    selling_price=float(selling_price),
                    min_stock=int(row_data.get('min_stock', 10) or 10),
                    status=1,
                )

                db.session.add(product)
                db.session.flush()

                quantity = int(row_data.get('quantity', 0) or 0)
                set_inventory_quantity(
                    product.product_id,
                    quantity,
                    change_type='import',
                    reason='Excel导入初始库存',
                )

                success_count += 1

            except Exception as e:
                errors.append(f'第 {row_num} 行：{str(e)}')
                error_count += 1

        db.session.commit()
        wb.close()
        return success_count, error_count, errors

    except Exception as e:
        db.session.rollback()
        return 0, 0, [f'导入失败：{str(e)}']


def get_products(page=1, per_page=20, search='', category_id=None, status=None):
    """
    获取商品列表

    Args:
        page: 页码
        per_page: 每页数量
        search: 搜索关键词
        category_id: 分类ID
        status: 状态

    Returns:
        dict: {total, pages, items}
    """
    query = Product.query.outerjoin(Category).outerjoin(Inventory)

    if search:
        query = query.filter(
            db.or_(
                Product.product_name.like(f'%{search}%'),
                Product.product_code.like(f'%{search}%'),
                Product.barcode.like(f'%{search}%'),
            )
        )

    if category_id:
        query = query.filter(Product.category_id == category_id)

    if status is not None:
        query = query.filter(Product.status == status)

    query = query.order_by(Product.product_id.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    items = []
    for product in pagination.items:
        inventory = product.inventory
        quantity = inventory.quantity if inventory else 0

        profit_rate = 0
        if product.selling_price > 0:
            profit_rate = ((product.selling_price - product.purchase_price) / product.selling_price * 100)

        if quantity <= 0:
            stock_status = 'out'
        elif quantity <= product.min_stock:
            stock_status = 'low'
        else:
            stock_status = 'normal'

        items.append({
            'product_id': product.product_id,
            'product_code': product.product_code,
            'barcode': product.barcode or '',
            'product_name': product.product_name,
            'category_name': product.category.category_name if product.category else '-',
            'unit': product.unit,
            'purchase_price': float(product.purchase_price),
            'selling_price': float(product.selling_price),
            'profit_rate': f'{profit_rate:.1f}%',
            'quantity': quantity,
            'min_stock': product.min_stock,
            'stock_status': stock_status,
            'status': product.status,
            'created_at': product.created_at.strftime('%Y-%m-%d') if product.created_at else '-',
        })

    return {
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'items': items,
    }


def create_product(data):
    """
    创建商品

    Args:
        data: 商品数据字典

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        if Product.query.filter_by(product_code=data['product_code']).first():
            return False, '商品编码已存在'

        barcode = data.get('barcode', '').strip()
        if barcode and Product.query.filter_by(barcode=barcode).first():
            return False, '条形码已存在'

        product = Product(
            product_code=data['product_code'],
            barcode=barcode if barcode else None,
            product_name=data['product_name'],
            category_id=data.get('category_id') or None,
            unit=data.get('unit', '件'),
            purchase_price=data.get('purchase_price', 0),
            selling_price=data['selling_price'],
            min_stock=data.get('min_stock', 10),
            status=1,
        )

        db.session.add(product)
        db.session.flush()

        set_inventory_quantity(
            product.product_id,
            data.get('quantity', 0),
            change_type='create',
            reason='商品新增初始库存',
        )
        db.session.commit()

        return True, '商品创建成功'
    except Exception as e:
        db.session.rollback()
        return False, f'创建失败：{str(e)}'


def update_product(product_id, data):
    """
    更新商品

    Args:
        product_id: 商品ID
        data: 更新数据

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        product = Product.query.get(product_id)
        if not product:
            return False, '商品不存在'

        if 'product_code' in data:
            existing = Product.query.filter_by(product_code=data['product_code']).first()
            if existing and existing.product_id != product_id:
                return False, '商品编码已存在'

        if 'barcode' in data:
            barcode = data['barcode'].strip() if data['barcode'] else ''
            if barcode:
                existing = Product.query.filter_by(barcode=barcode).first()
                if existing and existing.product_id != product_id:
                    return False, '条形码已存在'
                product.barcode = barcode
            else:
                product.barcode = None

        if 'product_code' in data:
            product.product_code = data['product_code']
        if 'product_name' in data:
            product.product_name = data['product_name']
        if 'category_id' in data:
            product.category_id = data['category_id'] or None
        if 'unit' in data:
            product.unit = data['unit']
        if 'purchase_price' in data:
            product.purchase_price = data['purchase_price']
        if 'selling_price' in data:
            product.selling_price = data['selling_price']
        if 'min_stock' in data:
            product.min_stock = data['min_stock']
        if 'status' in data:
            product.status = data['status']

        if 'quantity' in data:
            set_inventory_quantity(
                product_id,
                data['quantity'],
                change_type='adjust',
                reason='商品编辑调整库存',
            )

        db.session.commit()
        return True, '商品更新成功'
    except Exception as e:
        db.session.rollback()
        return False, f'更新失败：{str(e)}'


def offline_product(product_id):
    """下架商品（软删除）"""
    try:
        product = Product.query.get(product_id)
        if not product:
            return False, '商品不存在'

        product.status = 0
        db.session.commit()
        return True, '商品已下架'
    except Exception as e:
        db.session.rollback()
        return False, f'下架失败：{str(e)}'


def online_product(product_id):
    """上架商品"""
    try:
        product = Product.query.get(product_id)
        if not product:
            return False, '商品不存在'

        product.status = 1
        db.session.commit()
        return True, '商品已上架'
    except Exception as e:
        db.session.rollback()
        return False, f'上架失败：{str(e)}'


def delete_product(product_id):
    """删除商品（真删除）"""
    try:
        product = Product.query.get(product_id)
        if not product:
            return False, '商品不存在'

        inventory = Inventory.query.filter_by(product_id=product_id).first()
        if inventory:
            db.session.delete(inventory)

        db.session.delete(product)
        db.session.commit()
        return True, '商品已删除'
    except Exception as e:
        db.session.rollback()
        return False, f'删除失败：{str(e)}'


def get_categories():
    """获取所有分类"""
    categories = Category.query.filter_by(parent_id=0).order_by(Category.sort_order).all()
    return [{'id': c.category_id, 'name': c.category_name} for c in categories]
